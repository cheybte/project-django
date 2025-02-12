from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse
from .models import Wilaya, Moughataa, Commune, ProductType, Product, PointOfSale, Cart, ProductPrice, CartProduct

from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd




def import_wilayas(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('wilaya-list')

        try:
            df = pd.read_excel(excel_file)
            if not set(['code', 'name']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code' et 'name'.")
                return redirect('wilaya-list')
            Wilaya.objects.all().delete()
            for _, row in df.iterrows():
                Wilaya.objects.create(code=row['code'], name=row['name'])
            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('wilaya-list')
    return render(request, 'wilaya/import.html')
    
def import_moughataas(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('moughataa-list')

        try:
            df = pd.read_excel(excel_file)
            if not set(['code', 'label', 'wilaya_id']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code', 'label' et 'wilaya_id'.")
                return redirect('moughataa-list')
            Moughataa.objects.all().delete()
            for _, row in df.iterrows():
                Moughataa.objects.create(code=row['code'], label=row['label'], wilaya_id=row['wilaya_id'])
            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('moughataa-list')
    return render(request, 'moughataa/import.html')

def import_communes(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('commune-list')

        try:
            df = pd.read_excel(excel_file)
            if not set(['code', 'name', 'moughataa_id']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code', 'name' et 'moughataa_id'.")
                return redirect('commune-list')
            Commune.objects.all().delete()
            for _, row in df.iterrows():
                Commune.objects.create(code=row['code'], name=row['name'], moughataa_id=row['moughataa_id'])
            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('commune-list')
    return render(request, 'commune/import.html')

def import_carts(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('cart-list')

        try:
            df = pd.read_excel(excel_file)
            if not set(['code', 'name', 'description']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code', 'name' et 'description'.")
                return redirect('cart-list')
            Cart.objects.all().delete()
            for _, row in df.iterrows():
                Cart.objects.create(code=row['code'], name=row['name'], description=row['description'])
            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('cart-list')
    return render(request, 'cart/import.html')

def import_cartproducts(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('cartproduct-list')

        try:
            # Charger le fichier Excel
            df = pd.read_excel(excel_file)

            # Vérifier les colonnes nécessaires
            required_columns = ['product_code', 'cart_code', 'weight', 'date_from', 'date_to']
            if not set(required_columns).issubset(df.columns):
                messages.error(request, f"Le fichier Excel doit contenir les colonnes {', '.join(required_columns)}.")
                return redirect('cartproduct-list')

            # Supprimer toutes les entrées existantes (si nécessaire)
            CartProduct.objects.all().delete()

            # Parcourir chaque ligne du fichier pour importer les données
            for _, row in df.iterrows():
                # Valider les relations avec Product et Cart
                try:
                    product = Product.objects.get(code=row['product_code'])
                    cart = Cart.objects.get(code=row['cart_code'])
                except Product.DoesNotExist:
                    messages.error(request, f"Produit avec le code '{row['product_code']}' introuvable.")
                    continue
                except Cart.DoesNotExist:
                    messages.error(request, f"Panier avec le code '{row['cart_code']}' introuvable.")
                    continue

                # Créer un CartProduct
                CartProduct.objects.create(
                    product=product,
                    cart=cart,
                    weight=row['weight'],
                    date_from=row['date_from'],
                    date_to=row['date_to']
                )

            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

        return redirect('cartproduct-list')

    return render(request, 'cartproduct/import.html')

def import_products(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('product-import')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Veuillez importer un fichier Excel valide (extension .xls ou .xlsx).")
            return redirect('product-import')

        try:
            # Charger le fichier Excel
            df = pd.read_excel(excel_file)

            # Vérifier les colonnes nécessaires
            if not set(['code', 'name', 'description', 'unit_measure', 'product_type']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code', 'name', 'description', 'unit_measure', et 'product_type'.")
                return redirect('product-import')

            # Importer ou mettre à jour les données
            for _, row in df.iterrows():
                if pd.isnull(row['code']) or pd.isnull(row['name']) or pd.isnull(row['unit_measure']):
                    messages.error(request, "Le fichier contient des lignes avec des données manquantes.")
                    return redirect('product-import')

                # Récupérer ou créer le ProductType
                product_type, created = ProductType.objects.get_or_create(
                    label=row['product_type'],
                    defaults={
                        'description': 'Auto-created from import'  # Valeur par défaut si non existant
                    }
                )

                # Créer ou mettre à jour le produit
                Product.objects.update_or_create(
                    code=row['code'],
                    defaults={
                        'name': row['name'],
                        'description': row.get('description', ''),
                        'unit_measure': row['unit_measure'],
                        'product_type': product_type,
                    }
                )

            messages.success(request, "Les produits ont été importés avec succès.")

        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

        return redirect('product-list')

    return render(request, 'product/import.html')



def import_pointofsales(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('pointofsale-import')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Veuillez importer un fichier Excel valide (extension .xls ou .xlsx).")
            return redirect('pointofsale-import')

        try:
            # Charger le fichier Excel avec pandas
            df = pd.read_excel(excel_file)

            # Vérifier les colonnes nécessaires
            required_columns = {'code', 'type', 'gps_lat', 'gps_lon', 'commune_code'}
            if not required_columns.issubset(df.columns):
                messages.error(request, f"Le fichier Excel doit contenir les colonnes {', '.join(required_columns)}.")
                return redirect('pointofsale-import')

            # Importer ou mettre à jour les données
            for _, row in df.iterrows():
                if pd.isnull(row['code']) or pd.isnull(row['type']) or pd.isnull(row['commune_code']):
                    messages.error(request, "Le fichier contient des lignes avec des données manquantes.")
                    return redirect('pointofsale-import')

                try:
                    # Récupérer la commune associée
                    commune = Commune.objects.get(code=row['commune_code'])

                    # Créer ou mettre à jour le point de vente
                    PointOfSale.objects.update_or_create(
                        code=row['code'],
                        defaults={
                            'type': row['type'],
                            'gps_lat': row['gps_lat'],
                            'gps_lon': row['gps_lon'],
                            'commune': commune
                        }
                    )
                except Commune.DoesNotExist:
                    messages.error(request, f"Commune avec le code '{row['commune_code']}' introuvable.")
                    return redirect('pointofsale-import')

            messages.success(request, "Les points de vente ont été importés avec succès.")

        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

        return redirect('pointofsale-list')

    return render(request, 'pointofsale/import.html')


def import_productprices(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('productprice-import')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Veuillez importer un fichier Excel valide (extension .xls ou .xlsx).")
            return redirect('productprice-import')

        try:
            # Charger le fichier Excel
            df = pd.read_excel(excel_file)

            # Vérifier les colonnes nécessaires
            required_columns = ['product_code', 'point_of_sale_code', 'value', 'date_from', 'date_to']
            if not set(required_columns).issubset(df.columns):
                messages.error(request, f"Le fichier Excel doit contenir les colonnes : {', '.join(required_columns)}.")
                return redirect('productprice-import')

            # Importer ou mettre à jour les données
            for _, row in df.iterrows():
                # Vérifier les données obligatoires
                if pd.isnull(row['product_code']) or pd.isnull(row['point_of_sale_code']) or pd.isnull(row['value']):
                    messages.error(request, "Le fichier contient des lignes avec des données manquantes.")
                    return redirect('productprice-import')

                # Trouver le produit
                product = Product.objects.filter(code=row['product_code']).first()
                if not product:
                    messages.error(request, f"Le produit avec le code {row['product_code']} n'existe pas.")
                    continue

                # Trouver le point de vente
                point_of_sale = PointOfSale.objects.filter(code=row['point_of_sale_code']).first()
                if not point_of_sale:
                    messages.error(request, f"Le point de vente avec le code {row['point_of_sale_code']} n'existe pas.")
                    continue

                # Créer ou mettre à jour l'enregistrement ProductPrice
                ProductPrice.objects.create(
                    product=product,
                    point_of_sale=point_of_sale,
                    value=row['value'],
                    date_from=row['date_from'],
                    date_to=row['date_to']
                )

            messages.success(request, "Les prix des produits ont été importés avec succès.")

        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

        return redirect('productprice-list')

    return render(request, 'productprice/import.html')





def import_producttypes(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect('producttypes-import')

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Veuillez importer un fichier Excel valide (extension .xls ou .xlsx).")
            return redirect('producttypes-import')

        try:
            # Charger le fichier Excel
            df = pd.read_excel(excel_file)

            # Vérifier les colonnes nécessaires
            if not set(['code', 'label', 'description']).issubset(df.columns):
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code', 'label' et 'description'.")
                return redirect('producttypes-import')

            # Importer ou mettre à jour les données
            for _, row in df.iterrows():
                if pd.isnull(row['code']) or pd.isnull(row['label']):
                    messages.error(request, "Le fichier contient des lignes avec des données manquantes.")
                    return redirect('producttypes-import')

                ProductType.objects.update_or_create(
                    code=row['code'],
                    defaults={
                        'label': row['label'],
                        'description': row.get('description', ''),
                    }
                )

            messages.success(request, "Les ProductTypes ont été importés avec succès.")

        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

        return redirect('producttype-list')

    return render(request, 'producttype/import.html')





def import_data(request, model, expected_columns, success_url, template_name):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect(success_url)

        try:
            df = pd.read_excel(excel_file)
            if not set(expected_columns).issubset(df.columns):
                messages.error(request, f"Le fichier Excel doit contenir les colonnes {', '.join(expected_columns)}.")
                return redirect(success_url)
            model.objects.all().delete()
            for _, row in df.iterrows():
                model.objects.create(**{col: row[col] for col in expected_columns})
            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect(success_url)
    return render(request, template_name)


import openpyxl


def export_data(request, model, filename, fields):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = filename

    # Ajouter les en-têtes
    for idx, field in enumerate(fields, start=1):
        sheet.cell(row=1, column=idx, value=field)

    # Ajouter les données
    for row_idx, obj in enumerate(model.objects.all(), start=2):
        for col_idx, field in enumerate(fields, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=getattr(obj, field))

    workbook.save(response)
    return response

def export_wilayas(request):
    return export_data(
        request,
        model=Wilaya,
        filename='wilayas',
        fields=['code', 'name']
    )



from django.http import HttpResponse
import pandas as pd

def export_moughataas(request):
    # Récupérer les données de Moughataa
    moughataas = Moughataa.objects.all().values('code', 'label', 'wilaya__name')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(moughataas))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'code': 'Code',
        'label': 'Label',
        'wilaya__name': 'Wilaya'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="moughataas.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Moughataas')

    return response




def export_communes(request):
    # Récupérer les données de Commune
    communes = Commune.objects.all().values('code', 'name', 'moughataa__label', 'moughataa__wilaya__name')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(communes))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'code': 'Code',
        'name': 'Name',
        'moughataa__label': 'Moughataa',
        'moughataa__wilaya__name': 'Wilaya'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="communes.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Communes')

    return response






def export_carts(request):
    return export_data(
        request,
        model=Cart,
        filename='carts',
        fields=['code', 'name', 'description']  # Colonnes à inclure dans l'export
    )



from django.http import HttpResponse
import pandas as pd

def export_cartproducts(request):
    # Récupérer les données de CartProduct
    cartproducts = CartProduct.objects.all().values('cart__name', 'product__name', 'weight', 'date_from', 'date_to')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(cartproducts))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'cart__name': 'Cart Name',
        'product__name': 'Product Name',
        'weight': 'Weight',
        'date_from': 'Date From',
        'date_to': 'Date To'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cartproducts.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='CartProducts')

    return response




from django.http import HttpResponse
import pandas as pd

def export_products(request):
    # Récupérer les données de Product
    products = Product.objects.all().values('code', 'name', 'description', 'unit_measure', 'product_type__label')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(products))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'code': 'Code',
        'name': 'Name',
        'description': 'Description',
        'unit_measure': 'Unit Measure',
        'product_type__label': 'Product Type'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Products')

    return response




from django.http import HttpResponse
import pandas as pd

def export_pointofsales(request):
    # Récupérer les données de PointOfSale
    pointofsales = PointOfSale.objects.all().values('code', 'type', 'gps_lat', 'gps_lon', 'commune__name')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(pointofsales))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'code': 'Code',
        'type': 'Type',
        'gps_lat': 'GPS Latitude',
        'gps_lon': 'GPS Longitude',
        'commune__name': 'Commune'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="points_of_sale.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PointsOfSale')

    return response

    


def export_productprices(request):
    # Récupérer les données de ProductPrice
    productprices = ProductPrice.objects.all().values('value', 'date_from', 'date_to', 'product__name', 'point_of_sale__code')

    # Convertir les données en DataFrame pandas
    df = pd.DataFrame(list(productprices))

    # Renommer les colonnes pour un affichage plus clair
    df.rename(columns={
        'value': 'Value',
        'date_from': 'Date From',
        'date_to': 'Date To',
        'product__name': 'Product Name',
        'point_of_sale__code': 'Point of Sale Code'
    }, inplace=True)

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="product_prices.xlsx"'

    # Exporter le DataFrame vers Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ProductPrices')

    return response





def export_producttypes(request):
    return export_data(
        request,
        model=ProductType,
        filename='product_types',
        fields=['code', 'label', 'description']  # Colonnes exportées
    )


import matplotlib.pyplot as plt
from io import BytesIO
import base64



from django.db.models import Avg
from datetime import datetime
from django.shortcuts import render
from app.models import ProductType, Product, ProductPrice, CartProduct

from django.db.models import Avg
from django.shortcuts import render
from datetime import datetime
from app.models import ProductType, Product, ProductPrice, CartProduct

from calendar import month_name
def calculer_inpc(request):
    """
    Calcule l'Indice National des Prix à la Consommation (INPC)
    avec possibilité de sélectionner l'année et le mois.
    """
    # Année de base
    ANNEE_BASE = 2019

    # Récupérer l'année et le mois à partir de la requête, sinon utiliser l'année et le mois actuels
    annee_courante = request.GET.get('annee', datetime.now().year)
    mois_courant = request.GET.get('mois', datetime.now().month)

    try:
        annee_courante = int(annee_courante)
        mois_courant = int(mois_courant)
    except ValueError:
        annee_courante = datetime.now().year
        mois_courant = datetime.now().month

    # Récupérer tous les types de produits
    types_produits = ProductType.objects.all()

    # Dictionnaires pour stocker les résultats
    inpc_par_groupe = []
    inpc_global = {}
    inpc_tendances_mensuelles = {}  # Pour stocker les tendances mensuelles par groupe

    # Calculer l'INPC pour chaque type de produit
    for type_produit in types_produits:
        # Récupérer les produits de ce type
        produits = Product.objects.filter(product_type=type_produit)

        # Variables pour les totaux
        prix_total_base = 0
        prix_total_courant = 0
        poids_total = 0
        produits_calcules = 0

        # Liste pour stocker les tendances mensuelles
        tendances_mensuelles = []

        for produit in produits:
            # Récupérer les prix pour l'année de base
            prix_base = ProductPrice.objects.filter(
                product=produit,
                date_from__year=ANNEE_BASE,
                date_from__month=mois_courant
            ).order_by('-date_from').first()

            # Récupérer les prix pour l'année courante
            prix_courant = ProductPrice.objects.filter(
                product=produit,
                date_from__year=annee_courante,
                date_from__month=mois_courant
            ).order_by('-date_from').first()

            # Récupérer le poids du produit dans les paniers
            cart_products = CartProduct.objects.filter(
                product=produit,
                date_from__year__lte=annee_courante,
                date_to__year__gte=annee_courante
            )

            # Vérifier que tous les éléments nécessaires sont présents
            if prix_base and prix_courant and cart_products.exists():
                poids_moyen = cart_products.aggregate(Avg('weight'))['weight__avg'] or 0

                if poids_moyen > 0:
                    prix_total_base += prix_base.value * poids_moyen
                    prix_total_courant += prix_courant.value * poids_moyen
                    poids_total += poids_moyen
                    produits_calcules += 1

        # Calculer l'INPC pour ce groupe
        inpc_groupe = (prix_total_courant / prix_total_base * 100) if prix_total_base > 0 else 0

        # Ajouter les données du groupe si des produits sont calculés
        if produits_calcules > 0:
            inpc_par_groupe.append({
                'Année': annee_courante,
                'Mois': mois_courant,
                'Groupe': type_produit.label,
                'INPC': inpc_groupe,
                'Produits Calculés': produits_calcules
            })

        # Calculer les tendances mensuelles pour ce groupe
        tendances_mensuelles = []
        for mois in range(1, 13):
            prix_total_base_mois = 0
            prix_total_courant_mois = 0
            poids_total_mois = 0

            for produit in produits:
                prix_base_mois = ProductPrice.objects.filter(
                    product=produit,
                    date_from__year=ANNEE_BASE,
                    date_from__month=mois
                ).order_by('-date_from').first()

                prix_courant_mois = ProductPrice.objects.filter(
                    product=produit,
                    date_from__year=annee_courante,
                    date_from__month=mois
                ).order_by('-date_from').first()

                cart_products_mois = CartProduct.objects.filter(
                    product=produit,
                    date_from__year__lte=annee_courante,
                    date_to__year__gte=annee_courante
                )

                if prix_base_mois and prix_courant_mois and cart_products_mois.exists():
                    poids_moyen_mois = cart_products_mois.aggregate(Avg('weight'))['weight__avg'] or 0

                    if poids_moyen_mois > 0:
                        prix_total_base_mois += prix_base_mois.value * poids_moyen_mois
                        prix_total_courant_mois += prix_courant_mois.value * poids_moyen_mois
                        poids_total_mois += poids_moyen_mois

            inpc_groupe_mois = (prix_total_courant_mois / prix_total_base_mois * 100) if prix_total_base_mois > 0 else 0
            tendances_mensuelles.append(inpc_groupe_mois)

        inpc_tendances_mensuelles[type_produit.label] = tendances_mensuelles

    # Calculer l'INPC global
    inpc_total = sum(groupe['INPC'] for groupe in inpc_par_groupe) / len(inpc_par_groupe) if inpc_par_groupe else 0
    inpc_global[(annee_courante, mois_courant)] = inpc_total

    # Préparer les options d'années pour le formulaire
    annees_disponibles = sorted(set(
        ProductPrice.objects.values_list('date_from__year', flat=True).distinct()
    ))

    context = {
        'annee_base': ANNEE_BASE,
        'annee_courante': annee_courante,
        'mois_courant': mois_courant,
        'annees_disponibles': annees_disponibles,
        'inpc_par_groupe': inpc_par_groupe,
        'inpc_global': inpc_global,
        'inpc_tendances_mensuelles': inpc_tendances_mensuelles  # Données pour le graphique
    }
    

    return render(request, 'inpc.html', context)












from django.shortcuts import render
from django.db.models import Q
from .models import Product

def product_filter(request):
    # Récupérer les paramètres de filtrage depuis l'URL
    name = request.GET.get('name', '')
    code = request.GET.get('code', '')
    product_type = request.GET.get('product_type', '')

    # Filtrer les produits en fonction des paramètres
    products = Product.objects.all()

    if name:
        products = products.filter(name__icontains=name)
    if code:
        products = products.filter(code__icontains=code)
    if product_type:
        products = products.filter(product_type__label__icontains=product_type)

    # Passer les produits filtrés au template
    context = {
        'products': products,
    }
    return render(request, 'product/filter.html', context)



from django.shortcuts import render
from django.db.models import Q
from .models import ProductPrice

def productprice_filter(request):
    # Récupérer les paramètres de filtrage depuis l'URL
    product = request.GET.get('product', '')
    point_of_sale = request.GET.get('point_of_sale', '')
    value = request.GET.get('value', '')

    # Filtrer les prix des produits en fonction des paramètres
    productprices = ProductPrice.objects.all()

    if product:
        productprices = productprices.filter(product__name__icontains=product)
    if point_of_sale:
        productprices = productprices.filter(point_of_sale__name__icontains=point_of_sale)
    if value:
        productprices = productprices.filter(value=value)

    # Passer les prix des produits filtrés au template
    context = {
        'productprices': productprices,
    }
    return render(request, 'productprice/filter.html', context)






from django.http import JsonResponse
from django.db.models import Avg, Count
from .models import Product, ProductPrice, ProductType
import numpy as np
import random

def get_chart_data(request):
    # 📌 Récupération des prix moyens par mois depuis la base de données
    price_data = ProductPrice.objects.values('date_from').annotate(avg_price=Avg('value')).order_by('date_from')

    # 📌 Transformation des données pour le Line Chart
    dates = [str(entry['date_from']) for entry in price_data]  # Conversion des dates en chaînes
    avg_prices = [entry['avg_price'] for entry in price_data]

    line_chart_data = {
        'labels': dates,
        'datasets': [{
            'label': 'Prix Moyen',
            'data': avg_prices,
            'borderColor': 'rgba(75, 192, 192, 1)',
            'fill': False,
            'tension': 0.4  # Rend la courbe plus fluide
        }]
    }

    # 📌 Récupération de la répartition des produits par type (Pie Chart)
    product_types = ProductType.objects.annotate(product_count=Count('products'))
    pie_chart_data = {
        'labels': [pt.label for pt in product_types],
        'datasets': [{
            'data': [pt.product_count for pt in product_types],
            'backgroundColor': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF'],
        }]
    }

    # 📌 Récupération des prix moyens par type de produit (Bar Chart)
    avg_prices_by_type = ProductPrice.objects.values('product__product_type__label').annotate(avg_price=Avg('value'))
    bar_chart_data = {
        'labels': [entry['product__product_type__label'] for entry in avg_prices_by_type],
        'datasets': [{
            'label': 'Prix Moyen',
            'data': [entry['avg_price'] for entry in avg_prices_by_type],
            'backgroundColor': 'rgba(153, 102, 255, 0.6)',
        }]
    }

    # 📌 Génération d'une courbe fluide basée sur les données réelles (modèle sinusoïdal)
    if len(avg_prices) > 0:
        x_values = np.linspace(0, len(avg_prices) - 1, len(avg_prices))
        y_values = [price + random.uniform(-5, 5) for price in avg_prices]  # Ajout d'une variation aléatoire

        sinusoidal_chart_data = {
            'labels': dates,
            'datasets': [{
                'label': 'Évolution Continue de l\'INPC',
                'data': y_values,
                'borderColor': 'rgba(255, 99, 132, 1)',
                'fill': False,
                'tension': 0.4  # Rend la courbe plus lisse
            }]
        }
    else:
        sinusoidal_chart_data = {
            'labels': [],
            'datasets': []
        }

    # 📌 Retourner les données JSON pour être utilisées par Chart.js
    return JsonResponse({
        'line_chart_data': line_chart_data,
        'pie_chart_data': pie_chart_data,
        'bar_chart_data': bar_chart_data,
        'sinusoidal_chart_data': sinusoidal_chart_data,
    })






import numpy as np
import json
from scipy.interpolate import make_interp_spline
from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice

def get_inpc_chart_data(request):
    # Récupérer les prix quotidiens et calculer l'INPC
    prices = (
        ProductPrice.objects
        .values('date_from')
        .annotate(avg_price=Avg('value'))
        .order_by('date_from')
    )

    # Extraire les données
    dates = [str(p['date_from']) for p in prices]
    inpc_values = [p['avg_price'] for p in prices]

    # Interpolation pour lisser les données (éviter les graphes en "marches")
    if len(inpc_values) > 3:  # S'assurer qu'il y a assez de points pour interpoler
        x = np.arange(len(inpc_values))
        x_smooth = np.linspace(x.min(), x.max(), 300)  # 300 points pour une courbe lisse
        spline = make_interp_spline(x, inpc_values, k=3)  # Interpolation spline cubique
        inpc_smooth = spline(x_smooth)
    else:
        x_smooth, inpc_smooth = x, inpc_values  # Pas d'interpolation si trop peu de points

    # Préparer les données pour Chart.js
    chart_data = {
        'labels': dates,
        'datasets': [{
            'label': "Évolution INPC Continue",
            'data': inpc_smooth.tolist(),
            'borderColor': 'rgba(255, 99, 132, 1)',
            'fill': False,
            'tension': 0.4  # Rend la courbe lisse
        }]
    }

    return JsonResponse(chart_data)










# --- VUES POUR WILAYA ---
class WilayaCreateView(CreateView):
    model = Wilaya
    # Wilaya
    fields = ['code', 'name']
    template_name = 'wilaya/create.html'
    success_url = reverse_lazy('wilaya-list')


class WilayaListView(ListView):
    model = Wilaya
    template_name = 'wilaya/list.html'
    context_object_name = 'wilayas'


class WilayaUpdateView(UpdateView):
    model = Wilaya
    fields = ['code', 'name']
    template_name = 'wilaya/update.html'
    success_url = reverse_lazy('wilaya-list')


class WilayaDeleteView(DeleteView):
    model = Wilaya
    template_name = 'wilaya/delete.html'
    success_url = reverse_lazy('wilaya-list')


# --- VUES POUR MOUGHATAA ---
class MoughataaCreateView(CreateView):
    model = Moughataa
    fields = ['code', 'label', 'wilaya']
    template_name = 'moughataa/create.html'
    success_url = reverse_lazy('moughataa-list')


class MoughataaListView(ListView):
    model = Moughataa
    template_name = 'moughataa/list.html'
    context_object_name = 'moughataas'


class MoughataaUpdateView(UpdateView):
    model = Moughataa
    fields = ['code', 'label', 'wilaya']
    template_name = 'moughataa/update.html'
    success_url = reverse_lazy('moughataa-list')


class MoughataaDeleteView(DeleteView):
    model = Moughataa
    template_name = 'moughataa/delete.html'
    success_url = reverse_lazy('moughataa-list')


# --- VUES POUR COMMUNE ---
class CommuneListView(ListView):
    model = Commune
    template_name = 'commune/list.html'
    context_object_name = 'communes'


class CommuneCreateView(CreateView):
    model = Commune
    fields = ['code', 'name', 'moughataa']
    template_name = 'commune/create.html'
    success_url = reverse_lazy('commune-list')


class CommuneUpdateView(UpdateView):
    model = Commune
    fields = ['code', 'name', 'moughataa']
    template_name = 'commune/update.html'
    success_url = reverse_lazy('commune-list')


class CommuneDeleteView(DeleteView):
    model = Commune
    template_name = 'commune/delete.html'
    success_url = reverse_lazy('commune-list')


# --- VUES POUR PRODUCTTYPE ---
class ProductTypeListView(ListView):
    model = ProductType
    template_name = 'producttype/list.html'
    context_object_name = 'producttypes'


class ProductTypeCreateView(CreateView):
    model = ProductType
    fields = ['code', 'label', 'description']
    template_name = 'producttype/create.html'
    success_url = reverse_lazy('producttype-list')


class ProductTypeUpdateView(UpdateView):
    model = ProductType
    fields = ['code', 'label', 'description']
    template_name = 'producttype/update.html'
    success_url = reverse_lazy('producttype-list')


class ProductTypeDeleteView(DeleteView):
    model = ProductType
    template_name = 'producttype/delete.html'
    success_url = reverse_lazy('producttype-list')


# --- VUES POUR PRODUCT ---
class ProductListView(ListView):
    model = Product
    template_name = 'product/list.html'
    context_object_name = 'products'


class ProductCreateView(CreateView):
    model = Product
    fields = ['code', 'name', 'description', 'unit_measure', 'product_type']
    template_name = 'product/create.html'
    success_url = reverse_lazy('product-list')


class ProductUpdateView(UpdateView):
    model = Product
    fields = ['code', 'name', 'description', 'unit_measure', 'product_type']
    template_name = 'product/update.html'
    success_url = reverse_lazy('product-list')


class ProductDeleteView(DeleteView):
    model = Product
    template_name = 'product/delete.html'
    success_url = reverse_lazy('product-list')


# --- VUES POUR POINT OF SALE ---
class PointOfSaleListView(ListView):
    model = PointOfSale
    template_name = 'pointofsale/list.html'
    context_object_name = 'points_of_sale'


class PointOfSaleCreateView(CreateView):
    model = PointOfSale
    fields = ['code', 'type', 'gps_lat', 'gps_lon', 'commune']
    template_name = 'pointofsale/create.html'
    success_url = reverse_lazy('pointofsale-list')


class PointOfSaleUpdateView(UpdateView):
    model = PointOfSale
    fields = ['code', 'type', 'gps_lat', 'gps_lon', 'commune']
    template_name = 'pointofsale/update.html'
    success_url = reverse_lazy('pointofsale-list')


class PointOfSaleDeleteView(DeleteView):
    model = PointOfSale
    template_name = 'pointofsale/delete.html'
    success_url = reverse_lazy('pointofsale-list')


# --- VUES POUR CART ---
class CartListView(ListView):
    model = Cart
    template_name = 'cart/list.html'
    context_object_name = 'carts'


class CartCreateView(CreateView):
    model = Cart
    fields = ['code', 'name', 'description']
    template_name = 'cart/create.html'
    success_url = reverse_lazy('cart-list')


class CartUpdateView(UpdateView):
    model = Cart
    fields = ['code', 'name', 'description']
    template_name = 'cart/update.html'
    success_url = reverse_lazy('cart-list')


class CartDeleteView(DeleteView):
    model = Cart
    template_name = 'cart/delete.html'
    success_url = reverse_lazy('cart-list')

class ProductPriceListView(ListView):
    model = ProductPrice
    template_name = 'productprice/list.html'
    context_object_name = 'product_prices'


class ProductPriceCreateView(CreateView):
    model = ProductPrice
    fields = ['value', 'date_from', 'date_to', 'product', 'point_of_sale']
    template_name = 'productprice/create.html'
    success_url = reverse_lazy('productprice-list')


class ProductPriceUpdateView(UpdateView):
    model = ProductPrice
    fields = ['value', 'date_from', 'date_to', 'product', 'point_of_sale']
    template_name = 'productprice/update.html'
    success_url = reverse_lazy('productprice-list')


class ProductPriceDeleteView(DeleteView):
    model = ProductPrice
    template_name = 'productprice/delete.html'
    success_url = reverse_lazy('productprice-list')
    

# Liste des CartProducts
class CartProductListView(ListView):
    model = CartProduct
    template_name = 'cartproduct/list.html'
    context_object_name = 'cartproducts'

# Créer un CartProduct
class CartProductCreateView(CreateView):
    model = CartProduct
    fields = ['product', 'cart', 'weight', 'date_from', 'date_to']
    template_name = 'cartproduct/create.html'
    success_url = reverse_lazy('cartproduct-list')

# Modifier un CartProduct
class CartProductUpdateView(UpdateView):
    model = CartProduct
    fields = ['product', 'cart', 'weight', 'date_from', 'date_to']
    template_name = 'cartproduct/update.html'
    success_url = reverse_lazy('cartproduct-list')

# Supprimer un CartProduct
class CartProductDeleteView(DeleteView):
    model = CartProduct
    template_name = 'cartproduct/delete.html'
    success_url = reverse_lazy('cartproduct-list')

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render

@login_required




def home(request):
    return render(request, 'home.html')  # Rend le fichier 'home.html'